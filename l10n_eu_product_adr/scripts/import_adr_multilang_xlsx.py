# Copyright 2021 Stefan Rijnhart <stefan@opener.amsterdam>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl)
# Transforms the spreadsheet found at https://www.cepa.be/wp-content/uploads\
#    /ADR_2019_BijlageA_deel3_Tabel_A_EXCEL_FORMAAT.xlsx
# to adr.goods XML data.

import re
import sys
from collections import defaultdict

from lxml import etree
from openpyxl import load_workbook  # pylint: disable=W7936

skiprows = 3
columns = {
    0: "un_number",
    2: "name",
    5: "class_id",
    6: "classification_code",
    8: "label_ids",
    9: "provisions_3.3",
    10: "limited_quantity",
    12: "packing_instruction_ids",
    20: "transport_category",
}
index = {key: value for value, key in columns.items()}

# Articles containing dangerous goods
# (as opposed to dangerous goods themselves)
article_labels = {
    "3537": "2.1",
    "3538": "2.2",
    "3539": "2.3",
    "3540": "3",
    "3541": "4.1",
    "3542": "4.2",
    "3543": "4.3",
    "3544": "5.1",
    "3545": "5.2",
    "3546": "6.1",
    "3547": "8",
    "3548": "9",
}

valid_categories = [
    "0",
    "1",
    "2",
    "3",
    "4",
    "-",
    "CARRIAGE_PROHIBITED",
    "NOT_SUBJECT_TO_ADR",
]

valid_tunnel_codes = [
    "B",
    "B1000C",
    "B/D",
    "B/E",
    "C5000D",
    "C",
    "C/D",
    "C/E",
    "D",
    "D/E",
    "E",
    "-",
    "CARRIAGE_PROHIBITED",
    "NOT_SUBJECT_TO_ADR",
]

description_quirks = {
    "1040": ("pressure", "lower_pressure"),
    "2814": ("animal material", "animal_material"),
    "2900": ("animal material", "animal_material"),
    "3256": ("below", "lower_flashpoint"),
    "3257": ("filled at or below", "lower_fill_temperature"),
    "3373": ("animal material", "animal_material"),
}

valid_labels = [
    "1",
    "1.4",
    "1.5",
    "1.6",
    "2.1",
    "2.2",
    "2.3",
    "3",
    "4.1",
    "4.2",
    "4.3",
    "5.1",
    "5.2",
    "6.1",
    "6.2",
    "7A",
    "7B",
    "7C",
    "7E",
    "8",
    "9",
    "9A",
]

uom_map = {
    "g": "uom.product_uom_gram",
    "kg": "uom.product_uom_kgm",
    "l": "uom.product_uom_litre",
    "ml": "l10n_eu_product_adr.product_uom_mililiter",
}

# Keytypes
single = []
by_qty = []
full_key = []

transformers = {}


def transformer(func):
    """
    Decorator to add functions to the `transformers` dictionary

    Functions are added by their function name
    """
    transformers[func.__name__] = func
    return func


def parse_provisions_3_3(row):
    value = row[index["provisions_3.3"]]
    return re.findall(r"\S+", str(value) if value else "")


def parse_packing_instructions(row):
    """Special effort to expand ellipsis 'P112 (a), (b)  of (c)'"""
    value = row[index["packing_instruction_ids"]]
    value = (str(value) if value else "").replace(",", "")
    if "2.2.7" in value or "1.7" in value:
        # Special packing instructions for radioactive goods
        return []
    parts = re.findall(r"\S+", str(value) if value else "")
    last_code = False
    res = []
    for part in parts:
        if part == "of":  # conjunction
            continue
        match = re.match("([A-Z0-9]+)", part)
        if match:
            last_code = match.group(1)
        elif part.startswith("("):
            if not res[-1].endswith(")"):
                res[-1] += part
                continue
            part = last_code + part
        res.append(part)
    res = [part.replace("(", "").replace(")", "") for part in res]
    return res


def parse_limited_quantity(row, split=True):
    """
    :param split: when set, parse into float and Odoo uom. Otherwise, return
    text representation as is.
    """
    value = row[index["limited_quantity"]] or ""
    if not value or value == "0":
        return "" if not split else (False, False)
    value = value.strip()
    if not split:
        return value
    match = re.match(r"([0-9]+\.?[0-9]?)\s+([a-zA-Z]+)", value)
    if not match:
        if "BP 251" in value:  # known case
            return False, False
        else:
            raise ValueError(f"Cannot parse limited quantity: {value} ({row})")
    quantity, uom_name = match.groups()
    if uom_name.lower() not in uom_map:
        raise ValueError(f"Unknown uom {uom_name} in limited quantity {value} ({row})")
    return quantity, uom_map[uom_name.lower()]


def apply_description_quirk(row):
    un_number = parse_un_number(row)
    if un_number in description_quirks:
        substring, term = description_quirks[un_number]
        if substring in row[index["name"]]:
            return term
    return False


def get_xml_id(row):
    un_number = parse_un_number(row)
    if un_number in single:
        parts = [un_number]
    else:
        qty = parse_limited_quantity(row, split=False).replace(" ", "")
        if un_number in by_qty:
            parts = [un_number, qty]
        else:
            parts = [
                un_number,
                parse_classification_code(row),
                "/".join(parse_labels(row)),
                "/".join(parse_provisions_3_3(row)),
                qty,
                "/".join(parse_packing_instructions(row)),
                parse_transport_category(row)[0],
                apply_description_quirk(row),
            ]
    res = "_".join(part for part in parts if part).replace(".", "dot")
    return f"adr_goods_{res}"


def parse_un_number(row):
    return str(row[index["un_number"]]).strip().zfill(4)


@transformer
def un_number(record, value, row):
    xml_id = get_xml_id(row)
    record.attrib.update(
        {
            "id": xml_id,
            "model": "adr.goods",
        }
    )
    etree.SubElement(
        record, "field", attrib={"name": "un_number"}
    ).text = parse_un_number(row)


@transformer
def packing_instruction_ids(record, value, row):
    refs = []
    for instruction in parse_packing_instructions(row):
        refs.append(f"ref('adr_packing_instruction_{instruction}')")
    expression = f"[(6, 0, [{', '.join(refs)}])]"
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "packing_instruction_ids",
            "eval": expression,
        },
    )


@transformer
def limited_quantity(record, value, row):
    quantity, uom = parse_limited_quantity(row)
    if quantity:
        etree.SubElement(
            record, "field", attrib={"name": "limited_quantity"}
        ).text = quantity
        etree.SubElement(
            record,
            "field",
            attrib={
                "name": "limited_quantity_uom_id",
                "ref": uom,
            },
        )


@transformer
def name(record, value, row):
    value = value.strip().replace("\n", "")
    etree.SubElement(record, "field", attrib={"name": "name"}).text = value


@transformer
def class_id(record, value, row):
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "class_id",
            "ref": f"adr_class_{value.replace('.', '_')}",
        },
    )


def parse_classification_code(row):
    value = row[index["classification_code"]] or ""
    # In case multiple codes are possible, take the first (rare)
    code = value.split(" of ")[0]
    return code.replace(",", ".").strip()  # Typo in the sheet


@transformer
def classification_code(record, value, row):
    if value is None:  # Not defined for code 0190, for example
        return
    code = parse_classification_code(row)
    etree.SubElement(
        record, "field", attrib={"name": "classification_code"}
    ).text = code


def parse_transport_category(row):
    value = row[index["transport_category"]]
    if value is None:
        if "VERVOER VERBODEN" in str(row):
            # codes 0020, 0021
            category = "CARRIAGE_PROHIBITED"
            tunnel_restriction_code = "CARRIAGE_PROHIBITED"
        elif "NIET ONDERWORPEN AAN HET ADR" in str(row):
            category = "NOT_SUBJECT_TO_ADR"
            tunnel_restriction_code = "NOT_SUBJECT_TO_ADR"
        elif str(row[0]) in ["2071", "3363"]:  # known exceptions
            category = "-"
            tunnel_restriction_code = "-"
    else:
        match = re.search(r"(.*)\(([^\)]+)\)", value, re.DOTALL)
        if not match:
            raise ValueError(
                "Unknown value for transport code/tunnel restriction code: "
                "{value} ({row})"
            )
        category = match.groups()[0].strip()
        tunnel_restriction_code = match.groups()[1].strip()
    if "BP671" in category:
        # Special provision 671. Category depending on packing group
        # or else "2"
        category = "2"
    if category == "_":
        category = "-"
    if category not in valid_categories:
        raise ValueError(f"Invalid transport category {category} in cell value {value}")
    if tunnel_restriction_code not in valid_tunnel_codes:
        raise ValueError(
            f"Invalid tunnel restriction code {tunnel_restriction_code} "
            "in cell value {value}"
        )
    return category, tunnel_restriction_code


@transformer
def transport_category(record, value, row):
    category, tunnel_restriction_code = parse_transport_category(row)
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "transport_category",
        },
    ).text = category
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "tunnel_restriction_code",
        },
    ).text = tunnel_restriction_code


def parse_labels(row):
    value = row[index["label_ids"]]
    labels = [label.strip() for label in (str(value) if value else "").split("+")]
    res = []
    for label in labels:
        if not label or label == "GEEN":  # Meaning: none
            continue
        if any(
            val in label
            for val in (
                # Carriage prohibited
                "VERVOER VERBODEN",
                # Not subject to ADR
                "NIET ONDERWORPEN AAN HET ADR",
            )
        ):
            break
        if "5.2.2.1.12" in label:
            un_number = parse_un_number(row)
            if un_number in article_labels:
                label = article_labels[un_number]
        if label not in valid_labels + ["7X"]:
            raise ValueError(f"Invalid label {label} in cell value {value}")
        res.append(label)
    return res


@transformer
def label_ids(record, value, row):
    labels = parse_labels(row)
    if "7X" in labels:
        labels.remove("7X")
        # Pick any one
        labels += ["7A", "7B", "7C", "7E"]
    label_refs = []
    for label in labels:
        label_refs.append(f"ref('adr_label_{label.replace('.', '_')}')")
    expression = f"[(6, 0, [{', '.join(label_refs)}])]"
    etree.SubElement(
        record,
        "field",
        attrib={
            "name": "label_ids",
            "eval": expression,
        },
    )


def transform_row(root, row):
    record = etree.SubElement(root, "record")
    for index, field in columns.items():
        if field not in transformers:
            continue
        value = row[index]
        if isinstance(value, int | float):
            value = str(value)
        try:
            transformers[field](record, value, row)
        except (ValueError, AttributeError) as e:
            raise ValueError(f"Could not transform row {row}: {e}") from e
    return record.attrib["id"]


def populate_key_types(sheet):
    """Determine key type per un number"""
    count = 0
    code2rows = defaultdict(list)
    for row in sheet.iter_rows(values_only=True):
        count += 1
        if count <= skiprows:
            continue
        if row[0] is None:  # Emtpy rows
            continue
        un_number = parse_un_number(row)
        code2rows[un_number].append(row)
    count = 0
    for key, rows in code2rows.items():
        count += 1
        if count <= skiprows:
            continue
        if len(rows) == 1:
            single.append(key)
        else:
            qtys = [parse_limited_quantity(row, split=False) for row in rows]
            if len(set(qtys)) == len(qtys):
                by_qty.append(key)
            else:
                full_key.append(key)


def import_adr_multilang_xlsx(argv):
    root = etree.Element("odoo")
    sheet = load_workbook(argv[0]).active
    count = 0
    seen = set()
    duplicates = []

    populate_key_types(sheet)

    for row in sheet.iter_rows(values_only=True):
        count += 1
        if count <= skiprows:
            continue
        if row[0] is None:  # Emtpy rows
            continue
        xmlid = transform_row(root, row)
        if xmlid in seen:
            duplicates.append(xmlid)
        else:
            seen.add(xmlid)
    if duplicates:
        raise ValueError("Duplicate xml ids:\n" + "\n".join(duplicates))
    print(  # pylint: disable=W8116
        etree.tostring(
            root, pretty_print=True, xml_declaration=True, encoding="utf-8"
        ).decode("utf-8")
    )


if __name__ == "__main__":
    import_adr_multilang_xlsx(sys.argv[1:])
