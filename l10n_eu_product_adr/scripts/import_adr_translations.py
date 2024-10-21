# Copyright 2021 Stefan Rijnhart <stefan@opener.amsterdam>
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl)
# Imports the translations in a database in which the adr.goods have already
# been imported from the same spreadsheet that serves as the input to this
# script. This is an Odoo shell script. Run as follows:
#
# export ADR_FILE=/tmp/ADR_2019_BijlageA_deel3_Tabel_A_EXCEL_FORMAAT.xlsx
# cat import_adr_translations.py | odoo.py shell -d <DATABASE>

import importlib.util
from os import environ

from openpyxl import load_workbook  # pylint: disable=W7936

from odoo.tools.misc import file_path

pyfile = file_path("l10n_eu_product_adr/scripts/import_adr_multilang_xlsx.py")
spec = importlib.util.spec_from_file_location("import_adr_multilang_xlsx", pyfile)
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)
get_xml_id = mod.get_xml_id

env = self.env  # noqa
skiprows = 3
columns = {
    1: "nl_NL",
    3: "de_DE",
    4: "fr_FR",
}


def activate_languages():
    installed = []
    for code in columns.values():
        mods = env["ir.module.module"].search([("state", "=", "installed")])
        lang = (
            env["res.lang"]
            .with_context(active_test=False)
            .search([("code", "=", code)])
        )
        if not lang:
            raise ValueError(f"Language with code {code} not found in Odoo")
        if not lang.active:
            lang.active = True
            installed.append(code)
    if installed:
        mods._update_translations(filter_lang=installed)


def import_adr_translations():
    if not environ.get("ADR_FILE"):
        raise ValueError(
            "Please define the location of the Excel sheet in the environment "
            "variable ADR_FILE"
        )
    sheet = load_workbook(environ["ADR_FILE"]).active
    mod.populate_key_types(sheet)
    count = 0
    for row in sheet.iter_rows(values_only=True):
        count += 1
        if count <= skiprows:
            continue
        if row[0] is None:  # Emtpy rows
            continue
        xml_id = get_xml_id(row)
        record = env.ref(f"l10n_eu_product_adr.{xml_id}")
        for index, lang in columns.items():
            if row[index]:
                translation = row[index].strip().replace("\n", "")
                record.with_context(lang=lang).write({"name": translation})


activate_languages()
import_adr_translations()
env.cr.commit()
